"""
CatalogCraft — Flask Backend
Serves the single-page app; all heavy processing stays client-side (canvas / jsPDF).
The Flask backend:
  - Serves index.html + static assets
  - Provides /api/session  (GET/POST) to persist session JSON server-side
  - Provides /api/excel    (POST)     to parse uploaded Excel on the server
    and return JSON rows (same as SheetJS does client-side, kept as fallback)
"""

from flask import Flask, render_template, jsonify, request, session
from flask import send_from_directory
import json, os, io

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "catalogcraft-dev-secret-2024")

# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/session", methods=["GET"])
def get_session():
    return jsonify(session.get("data", {}))


@app.route("/api/session", methods=["POST"])
def save_session():
    data = request.get_json(silent=True) or {}
    session["data"] = data
    return jsonify({"ok": True})


@app.route("/api/session/reset", methods=["POST"])
def reset_session():
    session.pop("data", None)
    return jsonify({"ok": True})


@app.route('/assets/<path:filename>')
def serve_assets(filename):
    return send_from_directory('assets', filename)

@app.route('/api/predefined-assets', methods=['GET'])
def get_predefined():
    return jsonify({
        "logo": "assets/logo.jpeg",
        "templates": {
            "1": "assets/1.jpeg",
            "2": "assets/2.jpeg",
            "3": "assets/3.jpeg"
        }
    })


@app.route("/api/excel", methods=["POST"])
def parse_excel():
    """Optional server-side Excel parsing (fallback if client SheetJS fails)."""
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl not installed"}), 501

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        return jsonify({"headers": headers, "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
